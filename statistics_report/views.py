import io

from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse
from django.shortcuts import render
from django.utils.dateparse import parse_date
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import Marker
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework.authentication import SessionAuthentication
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response
from rest_framework_simplejwt.authentication import JWTAuthentication

from .services import build_statistics

WORKBOOK_TITLE = 'Sama al Majd'

# Matches the site's primary brand color (src/index.css --color-primary,
# mirrored in statistics_report/static/statistics_report/admin_stats.css
# --stat-primary), so the exported workbook reads as the same product as
# the admin statistics page.
HEADER_FILL = PatternFill(start_color='FF14B8A6', end_color='FF14B8A6', fill_type='solid')
HEADER_FONT = Font(color='FFFFFFFF', bold=True)
BANNER_FONT = Font(color='FF14B8A6', bold=True, size=16)
ZEBRA_FILL = PatternFill(start_color='FFEEF5FA', end_color='FFEEF5FA', fill_type='solid')
THIN_SIDE = Side(style='thin', color='FFC7D9E6')
CELL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
MAX_COLUMN_WIDTH = 60
# Minimum data rows (beyond the header) before a sheet gets a native chart —
# a 1-2 point chart (e.g. the 2-way plane/bus transport split) isn't useful.
MIN_ROWS_FOR_CHART = 3

# Brand palette (hex, no '#') — same values as HEADER_FILL/ZEBRA_FILL above
# and statistics_report/static/statistics_report/admin_stats.css, so pie
# slices and bar/line series read as the same product as the dashboard.
CHART_PALETTE = ['14B8A6', '1B6FA8', '1A9C56', 'D92D20', '0E8074', '124D78', '6B8399', '3B5166']


def _write_sheet(sheet, headers, rows, banner=None):
    """Writes an optional title banner, a header row, then one row per
    record, and styles the result: bold/colored header, bordered cells,
    zebra-striped data rows, content-fit column widths, and the header row
    frozen so it stays visible while scrolling. Returns the row number the
    header ended up on (1, or 2 if a banner was written).
    """
    header_row = 1
    if banner:
        sheet.append([banner] + [''] * (len(headers) - 1))
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
        banner_cell = sheet.cell(row=1, column=1)
        banner_cell.font = BANNER_FONT
        banner_cell.alignment = Alignment(horizontal='left', vertical='center')
        header_row = 2

    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    for cell in sheet[header_row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical='center')

    first_data_row = header_row + 1
    for row in sheet.iter_rows(min_row=header_row, max_row=sheet.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = CELL_BORDER
        if row[0].row >= first_data_row and (row[0].row - first_data_row) % 2 == 1:
            for cell in row:
                cell.fill = ZEBRA_FILL

    for col_index, header in enumerate(headers, start=1):
        column_letter = get_column_letter(col_index)
        max_length = len(str(header))
        for row_index in range(first_data_row, sheet.max_row + 1):
            value = sheet.cell(row=row_index, column=col_index).value
            max_length = max(max_length, len(str(value)) if value is not None else 0)
        sheet.column_dimensions[column_letter].width = min(max_length + 4, MAX_COLUMN_WIDTH)

    sheet.freeze_panes = f'A{first_data_row}'
    return header_row


def _build_chart(kind, title, data_sheet, header_row, data_row_count, *,
                  num_columns=2, x_title=None, y_title=None, color=None,
                  width=16, height=8):
    """Builds a styled native Excel chart (a real embedded chart object, not
    just data) reading from data_sheet's existing table — never returns one
    if there isn't enough data for it to be meaningful. The chart isn't
    anchored to any sheet yet; callers place it with worksheet.add_chart().
    """
    if data_row_count < MIN_ROWS_FOR_CHART:
        return None

    if kind == 'bar':
        chart = BarChart()
        chart.type = 'col'
        chart.gapWidth = 60
    elif kind == 'line':
        chart = LineChart()
    elif kind == 'pie':
        chart = PieChart()
    else:
        raise ValueError(f'Unknown chart kind: {kind!r}')

    chart.title = title
    chart.height = height
    chart.width = width
    chart.style = 2

    data_ref = Reference(
        data_sheet, min_col=num_columns, max_col=num_columns,
        min_row=header_row, max_row=header_row + data_row_count,
    )
    categories_ref = Reference(
        data_sheet, min_col=1, max_col=1,
        min_row=header_row + 1, max_row=header_row + data_row_count,
    )
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(categories_ref)

    series = chart.series[0]
    series_color = color or CHART_PALETTE[0]

    if kind == 'pie':
        chart.legend.position = 'r'
        series.dLbls = _data_labels(show_percent=True)
        series.data_points = [
            DataPoint(idx=i, spPr=None) for i in range(data_row_count)
        ]
        for point, point_color in zip(series.data_points, CHART_PALETTE):
            point.graphicalProperties.solidFill = point_color
    else:
        chart.legend = None
        if kind == 'line':
            series.graphicalProperties.line.solidFill = series_color
            series.graphicalProperties.line.width = 25000  # ~2pt, in EMU
            series.marker = Marker(symbol='circle', size=6)
            series.marker.graphicalProperties.solidFill = series_color
            series.marker.graphicalProperties.line.solidFill = series_color
            series.smooth = False
        else:
            series.graphicalProperties.solidFill = series_color
            series.dLbls = _data_labels(show_val=True)

        if x_title:
            chart.x_axis.title = x_title
        if y_title:
            chart.y_axis.title = y_title
        chart.x_axis.delete = False
        chart.y_axis.delete = False

        # With more than ~10 categories (e.g. one point per booking date),
        # every label rotated diagonally collides with its neighbours and
        # with the axis title below them — thin them out to a readable count.
        if data_row_count > 10:
            chart.x_axis.tickLblSkip = -(-data_row_count // 10)  # ceil division

    return chart


def _data_labels(*, show_val=False, show_percent=False):
    """A DataLabelList with every component explicitly set. Leaving
    show_cat_name/show_ser_name/show_legend_key as their openpyxl default
    (None) renders as "on" in Excel/LibreOffice, which is what produced
    labels like "Bookings Count, Gulf of Aqaba, 4" stacked on every point —
    so every flag not explicitly requested here is forced off.
    """
    labels = DataLabelList()
    labels.showVal = show_val
    labels.showPercent = show_percent
    labels.showCatName = False
    labels.showSerName = False
    labels.showLegendKey = False
    labels.showBubbleSize = False
    return labels


# Statistics endpoints are reached both from the JWT-authenticated React app
# and from the session-authenticated admin_statistics_page below, so both
# authentication classes are accepted here.
ADMIN_AUTHENTICATION_CLASSES = [JWTAuthentication, SessionAuthentication]


def _parse_date_params(request):
    errors = {}
    start_date = request.query_params.get('start_date')
    end_date = request.query_params.get('end_date')

    if start_date:
        parsed_start = parse_date(start_date)
        if parsed_start is None:
            errors['start_date'] = 'Invalid date, expected YYYY-MM-DD.'
        start_date = parsed_start

    if end_date:
        parsed_end = parse_date(end_date)
        if parsed_end is None:
            errors['end_date'] = 'Invalid date, expected YYYY-MM-DD.'
        end_date = parsed_end

    return start_date, end_date, errors


@api_view(['GET'])
@authentication_classes(ADMIN_AUTHENTICATION_CLASSES)
@permission_classes([IsAdminUser])
def statistics_view(request):

    start_date, end_date, errors = _parse_date_params(request)
    if errors:
        return Response(errors, status=400)

    stats = build_statistics(start_date, end_date)
    return Response(stats, status=200)


@api_view(['GET'])
@authentication_classes(ADMIN_AUTHENTICATION_CLASSES)
@permission_classes([IsAdminUser])
def export_statistics(request):

    start_date, end_date, errors = _parse_date_params(request)
    if errors:
        return Response(errors, status=400)

    stats = build_statistics(start_date, end_date)

    workbook = Workbook()
    workbook.properties.title = WORKBOOK_TITLE
    workbook.properties.creator = WORKBOOK_TITLE
    workbook.properties.subject = WORKBOOK_TITLE

    summary_sheet = workbook.active
    summary_sheet.title = 'Summary'
    _write_sheet(summary_sheet, ['Metric', 'Value'], [
        ['Start date', stats['start_date'] or 'All time'],
        ['End date', stats['end_date'] or 'All time'],
        ['Total bookings', stats['total_bookings']],
        ['New users', stats['new_users']],
    ], banner=WORKBOOK_TITLE)

    bookings_over_time_sheet = workbook.create_sheet('Bookings Over Time')
    bookings_over_time_rows = [[row['date'], row['bookings_count']] for row in stats['bookings_over_time']]
    bot_header_row = _write_sheet(bookings_over_time_sheet, ['Date', 'Bookings Count'], bookings_over_time_rows)
    bot_chart = _build_chart(
        'line', 'Bookings over time', bookings_over_time_sheet, bot_header_row, len(bookings_over_time_rows),
        x_title='Date', y_title='Bookings', color=CHART_PALETTE[0],
    )
    if bot_chart:
        bookings_over_time_sheet.add_chart(bot_chart, f'D{bot_header_row}')

    new_users_sheet = workbook.create_sheet('New Users Over Time')
    new_users_rows = [[row['date'], row['new_users_count']] for row in stats['new_users_over_time']]
    nu_header_row = _write_sheet(new_users_sheet, ['Date', 'New Users Count'], new_users_rows)
    nu_chart = _build_chart(
        'line', 'New users over time', new_users_sheet, nu_header_row, len(new_users_rows),
        x_title='Date', y_title='New users', color=CHART_PALETTE[1],
    )
    if nu_chart:
        new_users_sheet.add_chart(nu_chart, f'D{nu_header_row}')

    trips_sheet = workbook.create_sheet('Most Booked Trips')
    trips_rows = [[row['trip_name'], row['bookings_count']] for row in stats['most_booked_trips']]
    trips_header_row = _write_sheet(trips_sheet, ['Trip Name', 'Bookings Count'], trips_rows)
    trips_chart = _build_chart(
        'bar', 'Most booked trips', trips_sheet, trips_header_row, len(trips_rows),
        x_title='Trip', y_title='Bookings', color=CHART_PALETTE[0],
    )
    if trips_chart:
        trips_sheet.add_chart(trips_chart, f'D{trips_header_row}')

    destination_sheet = workbook.create_sheet('Bookings By Destination')
    destination_rows = [[row['place'], row['bookings_count']] for row in stats['bookings_by_destination']]
    dest_header_row = _write_sheet(destination_sheet, ['Destination', 'Bookings Count'], destination_rows)
    dest_chart = _build_chart(
        'pie', 'Bookings by destination', destination_sheet, dest_header_row, len(destination_rows),
    )
    if dest_chart:
        destination_sheet.add_chart(dest_chart, f'D{dest_header_row}')

    hotel_sheet = workbook.create_sheet('Bookings By Hotel')
    hotel_rows = [[row['hotel_name'], row['bookings_count']] for row in stats['bookings_by_hotel']]
    hotel_header_row = _write_sheet(hotel_sheet, ['Hotel', 'Bookings Count'], hotel_rows)
    hotel_chart = _build_chart(
        'bar', 'Bookings by hotel', hotel_sheet, hotel_header_row, len(hotel_rows),
        x_title='Hotel', y_title='Bookings', color=CHART_PALETTE[2],
    )
    if hotel_chart:
        hotel_sheet.add_chart(hotel_chart, f'D{hotel_header_row}')

    transport_sheet = workbook.create_sheet('Bookings By Transport Method')
    transport_rows = [[row['transport_method'], row['bookings_count']] for row in stats['bookings_by_transport']]
    transport_header_row = _write_sheet(transport_sheet, ['Transport Method', 'Bookings Count'], transport_rows)
    transport_chart = _build_chart(
        'bar', 'Bookings by transport method', transport_sheet, transport_header_row, len(transport_rows),
        x_title='Transport method', y_title='Bookings', color=CHART_PALETTE[3],
    )
    if transport_chart:
        transport_sheet.add_chart(transport_chart, f'D{transport_header_row}')

    # Overview: a 2x2 grid of the same charts, mirrored onto Summary so
    # everything (numbers + every chart) is visible on the sheet the
    # workbook opens to, without hunting through tabs. These reference the
    # exact same cells as the full-size charts above — no data is recomputed
    # or duplicated, only the chart objects are.
    overview_row = summary_sheet.max_row + 2
    overview_label_cell = summary_sheet.cell(row=overview_row, column=1, value='Overview')
    overview_label_cell.font = Font(bold=True, size=13, color='FF10233A')
    chart_row = overview_row + 1

    overview_charts = [
        _build_chart(
            'line', 'Bookings over time', bookings_over_time_sheet, bot_header_row, len(bookings_over_time_rows),
            x_title='Date', y_title='Bookings', color=CHART_PALETTE[0], width=13, height=7,
        ),
        _build_chart(
            'line', 'New users over time', new_users_sheet, nu_header_row, len(new_users_rows),
            x_title='Date', y_title='New users', color=CHART_PALETTE[1], width=13, height=7,
        ),
        _build_chart(
            'bar', 'Most booked trips', trips_sheet, trips_header_row, len(trips_rows),
            x_title='Trip', y_title='Bookings', color=CHART_PALETTE[0], width=13, height=7,
        ),
        _build_chart(
            'pie', 'Bookings by destination', destination_sheet, dest_header_row, len(destination_rows),
            width=13, height=7,
        ),
    ]
    grid_anchors = [f'A{chart_row}', f'J{chart_row}', f'A{chart_row + 17}', f'J{chart_row + 17}']
    for chart, anchor in zip(overview_charts, grid_anchors):
        if chart:
            summary_sheet.add_chart(chart, anchor)

    buffer = io.BytesIO()
    workbook.save(buffer)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="sama_al_majd_statistics.xlsx"'
    # Every export is freshly generated from the current database; without
    # this, a browser can serve a stale cached download for the same URL
    # (e.g. same date-filter query params) instead of re-requesting it.
    response['Cache-Control'] = 'no-store'
    return response


@staff_member_required
def admin_statistics_page(request):
    return render(request, 'statistics_report/admin_statistics.html')
