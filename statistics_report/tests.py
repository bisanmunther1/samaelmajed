import io
from datetime import date

from django.contrib.auth.models import User
from django.test import TestCase
from openpyxl import load_workbook
from rest_framework.test import APIClient

from hotels.models import Hotels
from profiles.models import Profile, Trip_per_user
from trip_features.models import Trip_features
from trips.models import Trips


class StatisticsTestBase(TestCase):

    def setUp(self):
        self.client = APIClient()

        self.admin_user = User.objects.create_user(
            username='admin', password='pass1234', is_staff=True,
        )
        self.regular_user = User.objects.create_user(
            username='regular', password='pass1234', is_staff=False,
        )

        admin_profile = Profile.objects.create(
            username='admin', email='admin@example.com', user=self.admin_user,
        )
        admin_profile.joined_at = date(2026, 1, 1)
        admin_profile.save()

        regular_profile = Profile.objects.create(
            username='regular', email='regular@example.com', user=self.regular_user,
        )
        regular_profile.joined_at = date(2026, 6, 1)
        regular_profile.save()

        self.trip_with_plane = Trips.objects.create(
            name='Cairo Trip', place='Cairo', price=100, rate=4.5, num=2,
            discount=0, desc='desc', type='City', available=True,
        )
        self.trip_with_bus = Trips.objects.create(
            name='Aswan Trip', place='Aswan', price=80, rate=4.0, num=1,
            discount=0, desc='desc', type='Nature', available=True,
        )

        Trip_features.objects.create(
            name=self.trip_with_plane,
            access_plane_1=True, access_plane_name_1='EgyptAir',
            access_plane_price_1=200, plane_start_at_1='2026-01-01T00:00:00Z',
        )
        Trip_features.objects.create(
            name=self.trip_with_bus,
            access_bus_1=True, access_bus_name_1='GoBus',
            access_bus_price_1=50, bus_start_at_1='2026-01-01T00:00:00Z',
        )

        self.nile_hotel = Hotels.objects.create(
            trip_name_id=self.trip_with_plane.name, name='Nile Hotel', price=50,
        )

        # In date range [2026-02-01, 2026-02-28]
        Trip_per_user.objects.create(
            username=admin_profile, trip=self.trip_with_plane, trip_date=date(2026, 2, 10),
            hotel=self.nile_hotel, hotel_reserve_date=date(2026, 2, 10), price=150,
        )
        Trip_per_user.objects.create(
            username=regular_profile, trip=self.trip_with_plane, trip_date=date(2026, 2, 15),
            hotel=self.nile_hotel, hotel_reserve_date=date(2026, 2, 15), price=150,
        )
        Trip_per_user.objects.create(
            username=regular_profile, trip=self.trip_with_bus, trip_date=date(2026, 2, 20),
            hotel=None, price=80,
        )
        # Outside the date range above
        Trip_per_user.objects.create(
            username=admin_profile, trip=self.trip_with_bus, trip_date=date(2026, 5, 1),
            hotel=None, price=80,
        )


class StatisticsPermissionTests(StatisticsTestBase):

    def test_anonymous_user_is_rejected(self):
        response = self.client.get('/statistics/')
        self.assertIn(response.status_code, (401, 403))

    def test_non_admin_user_is_forbidden(self):
        self.client.force_authenticate(user=self.regular_user)
        response = self.client.get('/statistics/')
        self.assertEqual(response.status_code, 403)

    def test_admin_user_is_allowed(self):
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get('/statistics/')
        self.assertEqual(response.status_code, 200)

    def test_export_non_admin_user_is_forbidden(self):
        self.client.force_authenticate(user=self.regular_user)
        response = self.client.get('/statistics/export/')
        self.assertEqual(response.status_code, 403)

    def test_export_admin_user_is_allowed(self):
        self.client.force_authenticate(user=self.admin_user)
        response = self.client.get('/statistics/export/')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('attachment', response['Content-Disposition'])


class StatisticsDataTests(StatisticsTestBase):

    def setUp(self):
        super().setUp()
        self.client.force_authenticate(user=self.admin_user)

    def test_totals_within_date_range(self):
        response = self.client.get('/statistics/', {
            'start_date': '2026-02-01', 'end_date': '2026-02-28',
        })
        self.assertEqual(response.status_code, 200)
        data = response.json()

        self.assertEqual(data['total_bookings'], 3)
        self.assertEqual(data['new_users'], 0)  # admin joined Jan, regular joined Jun; neither is in Feb

    def test_totals_without_date_range(self):
        response = self.client.get('/statistics/')
        self.assertEqual(response.status_code, 200)
        data = response.json()

        self.assertEqual(data['total_bookings'], 4)
        self.assertEqual(data['new_users'], 2)

    def test_most_booked_trips(self):
        response = self.client.get('/statistics/')
        data = response.json()

        counts = {row['trip_name']: row['bookings_count'] for row in data['most_booked_trips']}
        self.assertEqual(counts['Cairo Trip'], 2)
        self.assertEqual(counts['Aswan Trip'], 2)

    def test_bookings_by_destination(self):
        response = self.client.get('/statistics/')
        data = response.json()

        counts = {row['place']: row['bookings_count'] for row in data['bookings_by_destination']}
        self.assertEqual(counts['Cairo'], 2)
        self.assertEqual(counts['Aswan'], 2)

    def test_bookings_by_hotel(self):
        response = self.client.get('/statistics/')
        data = response.json()

        counts = {row['hotel_name']: row['bookings_count'] for row in data['bookings_by_hotel']}
        self.assertEqual(counts['Nile Hotel'], 2)
        self.assertNotIn(None, counts)
        self.assertNotIn('', counts)

    def test_bookings_by_transport(self):
        response = self.client.get('/statistics/')
        data = response.json()

        counts = {row['transport_method']: row['bookings_count'] for row in data['bookings_by_transport']}
        self.assertEqual(counts['plane'], 2)  # both Cairo Trip bookings
        self.assertEqual(counts['bus'], 2)    # both Aswan Trip bookings

    def test_invalid_date_returns_400(self):
        response = self.client.get('/statistics/', {'start_date': 'not-a-date'})
        self.assertEqual(response.status_code, 400)

    def test_export_invalid_date_returns_400(self):
        response = self.client.get('/statistics/export/', {'end_date': 'not-a-date'})
        self.assertEqual(response.status_code, 400)

    def test_export_produces_a_readable_workbook(self):
        response = self.client.get('/statistics/export/')
        workbook = load_workbook(io.BytesIO(response.content))
        self.assertEqual(
            workbook.sheetnames,
            [
                'Summary', 'Bookings Over Time', 'New Users Over Time',
                'Most Booked Trips', 'Bookings By Destination',
                'Bookings By Hotel', 'Bookings By Transport Method',
            ],
        )
        summary = workbook['Summary']
        # Row 1 is the brand title banner, row 2 is the header row.
        self.assertEqual(summary['A5'].value, 'Total bookings')
        self.assertEqual(summary['B5'].value, 4)

    def test_export_sheets_have_expected_column_headers(self):
        response = self.client.get('/statistics/export/')
        workbook = load_workbook(io.BytesIO(response.content))

        expected_headers = {
            'Summary': (2, ['Metric', 'Value']),
            'Bookings Over Time': (1, ['Date', 'Bookings Count']),
            'New Users Over Time': (1, ['Date', 'New Users Count']),
            'Most Booked Trips': (1, ['Trip Name', 'Bookings Count']),
            'Bookings By Destination': (1, ['Destination', 'Bookings Count']),
            'Bookings By Hotel': (1, ['Hotel', 'Bookings Count']),
            'Bookings By Transport Method': (1, ['Transport Method', 'Bookings Count']),
        }
        for sheet_name, (header_row, headers) in expected_headers.items():
            sheet = workbook[sheet_name]
            actual_headers = [cell.value for cell in sheet[header_row]][:len(headers)]
            self.assertEqual(actual_headers, headers, f'unexpected headers on {sheet_name!r}')

    def test_summary_sheet_has_brand_title_banner(self):
        response = self.client.get('/statistics/export/')
        workbook = load_workbook(io.BytesIO(response.content))
        self.assertEqual(workbook['Summary']['A1'].value, 'Sama al Majd')

    def test_workbook_properties_use_brand_name(self):
        response = self.client.get('/statistics/export/')
        workbook = load_workbook(io.BytesIO(response.content))
        self.assertEqual(workbook.properties.title, 'Sama al Majd')

    def test_export_sheets_have_one_row_per_record(self):
        response = self.client.get('/statistics/export/')
        workbook = load_workbook(io.BytesIO(response.content))

        trips_sheet = workbook['Most Booked Trips']
        # header row + one data row per distinct trip (Cairo Trip, Aswan Trip)
        self.assertEqual(trips_sheet.max_row, 3)
        rows = {trips_sheet.cell(row=r, column=1).value: trips_sheet.cell(row=r, column=2).value
                for r in range(2, trips_sheet.max_row + 1)}
        self.assertEqual(rows, {'Cairo Trip': 2, 'Aswan Trip': 2})

    def test_export_header_row_is_styled_and_frozen(self):
        response = self.client.get('/statistics/export/')
        workbook = load_workbook(io.BytesIO(response.content))

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            header_row = 2 if sheet_name == 'Summary' else 1
            header_cell = sheet.cell(row=header_row, column=1)
            self.assertTrue(header_cell.font.bold, f'{sheet_name!r} header is not bold')
            self.assertEqual(
                header_cell.fill.start_color.rgb, 'FF14B8A6',
                f'{sheet_name!r} header is not the site primary color',
            )
            self.assertEqual(
                sheet.freeze_panes, f'A{header_row + 1}',
                f'{sheet_name!r} header row is not frozen',
            )

    def test_export_data_rows_are_zebra_striped(self):
        response = self.client.get('/statistics/export/')
        workbook = load_workbook(io.BytesIO(response.content))

        # 4 distinct trip_date values in the fixture data, so rows 2-5 give
        # us two consecutive data rows to compare.
        sheet = workbook['Bookings Over Time']
        first_data_row_fill = sheet.cell(row=2, column=1).fill.start_color.rgb
        second_data_row_fill = sheet.cell(row=3, column=1).fill.start_color.rgb
        self.assertNotEqual(first_data_row_fill, second_data_row_fill)

    def test_export_column_widths_fit_content(self):
        response = self.client.get('/statistics/export/')
        workbook = load_workbook(io.BytesIO(response.content))

        trips_sheet = workbook['Most Booked Trips']
        # 'Trip Name' header (9 chars) is shorter than 'Cairo Trip' (10 chars),
        # so the fitted width must be driven by the longer data value, not
        # left at openpyxl's unset/default width.
        self.assertIsNotNone(trips_sheet.column_dimensions['A'].width)
        self.assertGreater(trips_sheet.column_dimensions['A'].width, len('Trip Name'))

    def test_export_sheets_with_enough_rows_have_native_charts(self):
        response = self.client.get('/statistics/export/')
        workbook = load_workbook(io.BytesIO(response.content))

        # 4 distinct trip_date values in the fixture data clears the
        # MIN_ROWS_FOR_CHART threshold, so this sheet must carry a real
        # embedded chart object, not just data.
        bookings_over_time_sheet = workbook['Bookings Over Time']
        self.assertGreaterEqual(len(bookings_over_time_sheet._charts), 1)

    def test_export_sheets_below_chart_threshold_have_no_chart(self):
        response = self.client.get('/statistics/export/')
        workbook = load_workbook(io.BytesIO(response.content))

        # Only 2 transport methods (plane/bus) in the fixture data — below
        # MIN_ROWS_FOR_CHART, so no chart should be added.
        transport_sheet = workbook['Bookings By Transport Method']
        self.assertEqual(len(transport_sheet._charts), 0)


class StatisticsAdminPageTests(StatisticsTestBase):

    def test_non_staff_is_redirected_to_login(self):
        self.client.login(username='regular', password='pass1234')
        response = self.client.get('/admin/statistics/')
        self.assertEqual(response.status_code, 302)

    def test_staff_can_view_the_page(self):
        self.client.login(username='admin', password='pass1234')
        response = self.client.get('/admin/statistics/')
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'statistics_report/admin_statistics.html')
